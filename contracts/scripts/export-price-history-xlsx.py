#!/usr/bin/env python3
"""
Export oracle mark history to a multi-sheet Excel workbook.

Usage (from repo root or contracts/):
  python3 scripts/export-price-history-xlsx.py
  python3 contracts/scripts/export-price-history-xlsx.py

Reads:
  frontend/public/data/price-history.json
Writes:
  frontend/public/data/price-history.xlsx
"""

from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Install openpyxl: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "frontend" / "public" / "data" / "price-history.json"
OUT = ROOT / "frontend" / "public" / "data" / "price-history.xlsx"


def main() -> None:
    data = json.loads(SRC.read_text())
    samples = data.get("samples") or []
    updated = data.get("updatedAt") or ""
    note = data.get("note") or ""

    by_m: dict[str, list] = {}
    for s in samples:
        by_m.setdefault(s["market"], []).append(s)
    for m in by_m:
        by_m[m].sort(key=lambda x: x["ts"])

    wb = Workbook()
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1A1A1A")
    title_font = Font(name="Arial", bold=True, size=14)
    subtitle_font = Font(name="Arial", size=10, color="555555")
    body_font = Font(name="Arial", size=10)
    blue_font = Font(name="Arial", size=10, color="0000FF")
    alt_fill = PatternFill("solid", fgColor="F7F7F7")
    currency_fmt = '$#,##0.0000;($#,##0.0000);"-"'
    pct_fmt = "0.00%"

    def style_header(ws, row: int, cols: int) -> None:
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def autosize(ws, min_w: int = 10, max_w: int = 48) -> None:
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            length = 0
            for cell in col:
                if cell.value is not None:
                    length = max(length, min(max_w, len(str(cell.value)) + 2))
            ws.column_dimensions[letter].width = max(min_w, length)

    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "PEPT Trade — Oracle price history"
    ws["A1"].font = title_font
    ws["A2"] = (
        f"Exported {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} · "
        f"source updatedAt={updated}"
    )
    ws["A2"].font = subtitle_font
    ws["A3"] = note
    ws["A3"].font = subtitle_font
    ws.merge_cells("A3:I3")

    ws["A5"] = "Storage note"
    ws["A5"].font = Font(name="Arial", bold=True, size=10)
    ws["A6"] = (
        "NOT in Supabase. Historical marks live in git JSON: "
        "frontend/public/data/price-history.json (mirror: contracts/data/price-history.json). "
        "Appended by GitHub Actions after each oracle pushPrice. "
        "Per-vendor offers are in glp1-last-scrape.json (current scrape only + previous for deltas)."
    )
    ws["A6"].font = subtitle_font
    ws["A6"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A6:I6")
    ws.row_dimensions[6].height = 48

    headers = [
        "Market",
        "Samples",
        "First (UTC)",
        "Last (UTC)",
        "First $/mg",
        "Last $/mg",
        "Min $/mg",
        "Max $/mg",
        "Change %",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=8, column=i, value=h)
    style_header(ws, 8, len(headers))

    markets_order = ["SEMA-PERP", "TIRZ-PERP", "RETA-PERP", "GLP1-IDX-PERP"]
    for m in sorted(by_m.keys()):
        if m not in markets_order:
            markets_order.append(m)

    r = 9
    for m in markets_order:
        arr = by_m.get(m) or []
        if not arr:
            continue
        prices = [x["price"] for x in arr if isinstance(x.get("price"), (int, float))]
        first, last = arr[0], arr[-1]
        ws.cell(row=r, column=1, value=m).font = body_font
        ws.cell(row=r, column=2, value=len(arr)).font = blue_font
        ws.cell(
            row=r,
            column=3,
            value=datetime.fromtimestamp(first["ts"], tz=timezone.utc).strftime("%Y-%m-%d %H:%M"),
        ).font = body_font
        ws.cell(
            row=r,
            column=4,
            value=datetime.fromtimestamp(last["ts"], tz=timezone.utc).strftime("%Y-%m-%d %H:%M"),
        ).font = body_font
        for col, val in (
            (5, first["price"]),
            (6, last["price"]),
            (7, min(prices) if prices else None),
            (8, max(prices) if prices else None),
        ):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = blue_font
            cell.number_format = currency_fmt
        # Prefer computed value so file opens without Excel recalc
        change = ((last["price"] - first["price"]) / first["price"]) if first["price"] else None
        cell = ws.cell(row=r, column=9, value=change)
        cell.font = body_font
        cell.number_format = pct_fmt
        if r % 2 == 0:
            for c in range(1, 10):
                ws.cell(row=r, column=c).fill = alt_fill
        r += 1

    total_row = r + 1
    ws.cell(row=total_row, column=1, value="Total samples").font = Font(
        name="Arial", bold=True, size=10
    )
    ws.cell(row=total_row, column=2, value=len(samples)).font = Font(
        name="Arial", bold=True, size=10
    )

    autosize(ws)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    # All Samples
    ws_all = wb.create_sheet("All Samples")
    all_headers = [
        "Market",
        "Timestamp (unix)",
        "Datetime UTC",
        "Price $/mg",
        "Source",
        "Tx hash",
    ]
    for i, h in enumerate(all_headers, 1):
        ws_all.cell(row=1, column=i, value=h)
    style_header(ws_all, 1, len(all_headers))

    sorted_all = sorted(samples, key=lambda x: (x["ts"], x["market"]))
    for i, s in enumerate(sorted_all, 2):
        ws_all.cell(row=i, column=1, value=s["market"]).font = body_font
        ws_all.cell(row=i, column=2, value=s["ts"]).font = blue_font
        dt = datetime.fromtimestamp(s["ts"], tz=timezone.utc)
        ws_all.cell(row=i, column=3, value=dt.replace(tzinfo=None))
        ws_all.cell(row=i, column=3).number_format = "yyyy-mm-dd hh:mm"
        ws_all.cell(row=i, column=3).font = body_font
        pc = ws_all.cell(row=i, column=4, value=s["price"])
        pc.font = blue_font
        pc.number_format = currency_fmt
        ws_all.cell(row=i, column=5, value=(s.get("source") or "")[:200]).font = body_font
        ws_all.cell(row=i, column=6, value=s.get("txHash") or "").font = body_font
        if i % 2 == 0:
            for c in range(1, 7):
                ws_all.cell(row=i, column=c).fill = alt_fill

    autosize(ws_all)
    ws_all.column_dimensions["C"].width = 18
    ws_all.column_dimensions["E"].width = 40
    ws_all.column_dimensions["F"].width = 20
    ws_all.auto_filter.ref = f"A1:F{len(sorted_all) + 1}"
    ws_all.freeze_panes = "A2"

    for m in ["SEMA-PERP", "TIRZ-PERP", "RETA-PERP", "GLP1-IDX-PERP"]:
        arr = by_m.get(m) or []
        ws_m = wb.create_sheet(m)
        for i, h in enumerate(
            ["Datetime UTC", "Price $/mg", "Source", "Tx hash", "Timestamp"], 1
        ):
            ws_m.cell(row=1, column=i, value=h)
        style_header(ws_m, 1, 5)
        for i, s in enumerate(arr, 2):
            dt = datetime.fromtimestamp(s["ts"], tz=timezone.utc)
            ws_m.cell(row=i, column=1, value=dt.replace(tzinfo=None))
            ws_m.cell(row=i, column=1).number_format = "yyyy-mm-dd hh:mm"
            ws_m.cell(row=i, column=1).font = body_font
            pc = ws_m.cell(row=i, column=2, value=s["price"])
            pc.font = blue_font
            pc.number_format = currency_fmt
            ws_m.cell(row=i, column=3, value=(s.get("source") or "")[:180]).font = body_font
            ws_m.cell(row=i, column=4, value=s.get("txHash") or "").font = body_font
            ws_m.cell(row=i, column=5, value=s["ts"]).font = body_font
        if len(arr) >= 2:
            chart = LineChart()
            chart.title = f"{m} $/mg"
            chart.style = 10
            chart.y_axis.title = "$/mg"
            chart.x_axis.title = "Sample"
            chart.height = 10
            chart.width = 18
            data_ref = Reference(ws_m, min_col=2, min_row=1, max_row=len(arr) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            ws_m.add_chart(chart, "G2")
        autosize(ws_m)
        ws_m.column_dimensions["A"].width = 18
        ws_m.column_dimensions["C"].width = 40
        ws_m.freeze_panes = "A2"
        if arr:
            ws_m.auto_filter.ref = f"A1:E{len(arr) + 1}"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT} ({OUT.stat().st_size} bytes) samples={len(samples)}")


if __name__ == "__main__":
    main()
